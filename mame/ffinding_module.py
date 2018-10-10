'''
Feature Finding Module
@author: Jamie Nunez
(C) 2018 - Pacific Northwest National Laboratory

Data processing and feature downselection methods, as described in:
Advancing Standards-Free Methods for the Identification of Small Molecules in Complex Samples
Nunez et al.
'''

#%% Imports
from bisect import bisect_left
import numpy as np
from openpyxl import load_workbook

import formula_module as fmod

#%% Globals

# Globals that don't change during optimization
REPS = 3
MODES = 2
MAX_FTICR_ERROR = 15  # Will need to add shift as well

# Parameters that do change during optimization.
# * = used only in ffinding, ** = used only in scoring
PARAMS = [1000, # IMS-MS, min int
          6, # IMS-MS, max mass error
          3, # IMS-MS, min samples
          1, # IMS-MS, max blanks
          5, # IMS-MS, max CCS error
          2123, # IMS-MS, int cutoff, positive
          2174, # IMS-MS, int cutoff, negative
          1, # FT-ICR, min int
          1.5, # FT-ICR, max mass error
          1, # FT-ICR, min samples
          0, # FT-ICR, max blanks
          3357.7, # FT-ICR, int cutoff, positive
          109.7, # FT-ICR, int cutoff, negative
          200, # High mass cutoff
          6 # Min score to be included
          ]


#%% Misc. Helper Functions
def column_string(n):
    '''
    Converts numbers to their corresponding column name in Excel.
    
    Parameters
    ----------
    n: int
        Column number
    
    Returns
    ----------
    s: string
        Excel column string
        
    Examples
    ----------
    >>> column_string(5)
    'E'
    
    >>> column_string(31)
    'AE'
    '''
    
    div = n
    s = ''
    while div > 0:
        module = (div - 1) % 26
        s = chr(65 + module) + s
        div=int((div - module) / 26)
    return s


def calc_error(val1, val2, absolute=False):
    '''
    Calculates the error between two values.
    Equation used: (val2 - val1) / val2
    
    Parameters
    ----------
    val1: float
        First number, taken to be the "new" value.
    val2: float
        Second number, taken to be the "true" or "accepted" value
    absolute: boolean (optional)
        If true, returns the absolute value of the error. Default value is False.
        
    Returns
    ----------
    e: float
        Error between val1 and val2
    '''
    if val1 is None or val2 is None:
        return None
    if absolute:
        return abs(val2 - val1) / val2
    else:
        return (val2 - val1) / val2


def ppm_error(val1, val2, absolute=False):
    '''
    Calculates the error (in ppm) between two values.
    Equation used: (val2 - val1) / val2 * 1e6
    
    Parameters
    ----------
    val1: float
        First number, taken to be the "new" value.
    val2: float
        Second number, taken to be the "true" or "accepted" value
    absolute: boolean (optional)
        If true, returns the absolute value of the error. Default value is False.
        
    Returns
    ----------
    e: float
        Error between val1 and val2 in ppm
    '''
    return calc_error(val1, val2, absolute=absolute) * 1e6


def percent_error(val1, val2, absolute=False):
    '''
    Calculates the error (in percent) between two values.
    Equation used: (val2 - val1) / val2 * 1e2
    
    Parameters
    ----------
    val1: float
        First number, taken to be the "new" value.
    val2: float
        Second number, taken to be the "true" or "accepted" value
    absolute: boolean (optional)
        If true, returns the absolute value of the error. Default value is False.
        
    Returns
    ----------
    e: float
        Error between val1 and val2 in percent
    '''
    return calc_error(val1, val2, absolute=absolute) * 1e2


def find_closest_sorted(l, value):
    '''
    Within a sorted (smallest to largest) list, l, finds the number
    closest to a given value.
    If two numbers are equally close, returns the smallest number.
    '''
    pos = bisect_left(l, value)
    if pos == 0:
        return l[0]
    if pos == len(l):
        return l[-1]
    before = l[pos - 1]
    after = l[pos]
    if after - value < value - before:
       return after
    else:
       return before


def find_closest(l, val):
    '''
    Within a list, l, finds the number closest to a given value.
    If two numbers are equally close, returns the number that comes
    first in the list.
    '''
    return min(l, key=lambda x:abs(x-val))


def in_list(l, val, e=PARAMS[2]/1000000., sorted_list=False):
    '''
    Checks if an element is within a given list within a given error.
    If it is found to be in the given list, returns the value closest
    to it.
    
    Parameters
    ----------
    l: list
        Elements to search. Can be sorted or unsorted.
    val: float
        Value to search for within list
    e: float
        Allowable error to count val as equivalent to an element in the
        list.
    sorted_list: boolean (optional)
        Pass in True if the list is sorted.
        Can make this function significantly faster with large lists.
        
    Returns
    ----------
    res: number, same type as elements in given list
        Closest element within the list if it is with e of the given
        val.
        If no elements are close enough to val, returns None.
    '''
    
    # Exit for invalid list
    if len(l) == 0:
        return None
        
    # Find closest mass in list
    if sorted_list:  # Achieve speed up using binary search
        closest = find_closest_sorted(l, val)
    else:
        closest = find_closest(l, val)
    if np.isclose(val, closest, atol=e):
        return closest
    return None


def present_in_mixes(feature_index, info, sample_num):
    i = feature_index
    l = [j for j in range(sample_num) if info[i, -sample_num + j] == 1]  # 1 symbolizes passed
    return l


def open_library(path='../data/', fname='SuspectLibrary_FromToxCast'):
    '''
    Open a spreadsheet and return it and its first sheet.
    '''
    wb = load_workbook(path + fname + '.xlsx')
    sheet = wb.worksheets[0]
    return wb, sheet


# Scrape info needed from lib. There's better/faster ways to do this.
def get_library_info(path='../data/', fname='SuspectLibrary_FromToxCast',
                     formula_col='H', mass_col='I', ccs_cols=['J', 'K', 'L'],
                     unique_col='M', present_col_num=14):
    _, sheet = open_library(path=path, fname=fname) 
    cpd_masses = []
    cpd_formulas = []
    cpd_ccs_vals = []
    cpd_present = []
    cpd_unique = []
    for i in range(2, sheet.max_row + 1):
        cpd_formulas.append(sheet[formula_col + str(i)].value)
        cpd_masses.append(float(sheet[mass_col + str(i)].value))
        ccs_vals = []
        for col in ccs_cols:
            ccs_vals.append(sheet[col + str(i)].value)
        cpd_ccs_vals.append(ccs_vals)
        cpd_unique.append(sheet[unique_col + str(i)].value) 
        present = []
        for j in range(10):
            present_col = column_string(present_col_num + j)
            present.append(sheet[present_col + str(i)].value)
        cpd_present.append(present)
    cpd_masses = np.array(cpd_masses, dtype=np.float64)
    cpd_ccs_vals = np.array(cpd_ccs_vals, dtype=np.float64)
    cpd_present = np.array(cpd_present, dtype=np.bool)
    cpd_unique = np.array(cpd_unique, dtype=np.float32)
    return cpd_formulas, cpd_masses, cpd_ccs_vals, cpd_present, cpd_unique


# Returns suggested mass shift for a given set of data based on how it matches
# to the suspect library. Shift given in ppm. Masses given should be in the
# form of lists.
def calc_mass_shift(adduct_list, mass_error, library_masses, feature_masses):
    
    # Generate adduct masses
    unique_masses = np.array(list(set(library_masses)), dtype=np.float64, ndmin=2)
    
    adduct_masses = (unique_masses + adduct_list[0]).T
    for z in adduct_list[1:]: 
        adduct_masses = np.vstack((adduct_masses, (unique_masses + z).T))
    adduct_masses = list(set(adduct_masses.T.tolist()[0]))
    
    # Prepare feature mass lists - ensure there are no repeats and they are sorted
    feature_masses = list(set(feature_masses))
    feature_masses.sort()
    
    # Calculate all mass errors
    mass_errors = []
    for mz in adduct_masses:
        mz_in_list = in_list(feature_masses, mz, e=mass_error, sorted_list=True)
        
        # Passed, now record information for matched feature
        if mz_in_list is not None:
            mass_errors.append(ppm_error(mz, mz_in_list))

    if len(mass_errors) == 0:
        return 0
    # Average of mass errors is the suggested mass shift
    return -np.nanmean(mass_errors)


#%% IMSMS

# Pull data from file
def _gen_imsms_data(fname, total_mixes, start_col=16):
    # Mix no: zero-based
    # File Settings (assumed to always be the same)
    usecols = [0, 7, 5]
    col_per_mix = MODES * REPS
    end_col = start_col + (total_mixes + 1) * col_per_mix
    usecols.extend(range(start_col, end_col))
    usecols = tuple(usecols)
    d = ','
    sr = 5    
    return np.loadtxt(fname, usecols=usecols, delimiter=d, skiprows=sr)


# Pull positive and negative data from files
def gen_imsms_data(fname='../data/IMSMS_%s_Features.csv', total_mixes=10):
    pos_features = _gen_imsms_data(fname % 'Pos', total_mixes)
    neg_features = _gen_imsms_data(fname % 'Neg', total_mixes)
    return pos_features, neg_features


# Calculate and apply a shift to all mass values in MS data
def shift_imsms_mass(pos, neg, library_masses):
    adducts = [1.0078246, 22.9898, -1.0078246]  # ['H', 'Na', 'De']
    pos_shift, neg_shift = np.nan, np.nan
    if PARAMS[12] == 1:
        pos_shift = calc_mass_shift(adducts[:2], PARAMS[2], library_masses, pos[:, 1])
        pos[:, 1] += pos_shift / 1000000.
        neg_shift = calc_mass_shift(adducts[2:], PARAMS[2], library_masses, neg[:, 1])
        neg[:, 1] += neg_shift / 1000000.
    return pos, neg, [pos_shift, neg_shift]


# Remove noise/contamination from MS data
def _downselect_imsms_features(features, min_int):

    # Count times seen in samples and blanks
    thresh = np.array(features > min_int, dtype=np.int)
    counts = np.zeros((features.shape[0], MODES * 2))
    int_col = 3  # This is where the sample intensities begin
    for j in range(MODES * 2):
        counts[:, j] = np.sum(thresh[:, int_col: int_col + REPS], axis=1)
        int_col += REPS
    
    # Cut out sample counts below PARAMS[3]. 0 to fail. 1 to pass.
    sample_counts = np.array(counts[:, :MODES] >= PARAMS[3], dtype=np.int)
    blank_counts = np.array(counts[:, MODES:] <= PARAMS[4], dtype=np.int)
    maxes = np.max(sample_counts + blank_counts, axis=1)
    
    features = np.vstack((features.T, sample_counts.T)).T
    features = np.delete(features, np.where(maxes < 2)[0], axis=0)    
    
    # Break into subsets
    info_cols = range(3)
    info_cols.extend(range(-MODES, 0))
    feature_info = features[:, info_cols]
    intensties = features[:, 3: 3 + MODES * REPS]
    
    return feature_info, intensties


# Remove noise/contamination from MS data (wrapper for pos and neg mode)
def downselect_imsms_features(pos_feat, neg_feat):
    pos_inf, pos_int = _downselect_imsms_features(pos_feat, PARAMS[0])
    neg_inf, neg_int = _downselect_imsms_features(neg_feat, PARAMS[1])
    pos_masses = list(set(pos_inf[:, 1]))  # Don't want repeat masses
    neg_masses = list(set(neg_inf[:, 1]))
    pos_int[pos_int == 0.001] = 0
    neg_int[neg_int == 0.001] = 0
    return pos_inf, pos_int, neg_inf, neg_int, pos_masses, neg_masses


# Return cols to use for a given mix number.
def get_usecols(mix_no, total_mixes=10, col_per_mix=MODES*REPS):
    col_per_mix = REPS * MODES
    usecols = range(3)
    start = 3 + mix_no * col_per_mix
    usecols.extend(range(start, start + col_per_mix))  # Sample intensitiy columns  
    start = 3 + total_mixes * col_per_mix
    usecols.extend(range(start, start + col_per_mix))  # Blank intensity columns
    return usecols


# Get all IMS-MS evidence for a given library entry's mass and CCS
def get_imsms_evidence(mz, ccs, mass_list, sample_info, sample_int):
    feature_info = []
    mz_in_list = in_list(mass_list, mz)
    while mz_in_list is not None:

        # Add features with this mass to the list of compounds
        id_matches = np.where(sample_info[:, 1] == mz_in_list)[0]

        l = []
        for ii in id_matches:
            
            l = present_in_mixes(ii, sample_info, 2)

            # Write each match
            for ll in l:
                if ll % 2 == 0:
                    avg_int = np.nanmean(sample_int[ii, :3])  # APPI
                else:
                    avg_int = np.nanmean(sample_int[ii, 3:])  # ESI

                # Start scoring
                this_feature_info = [0] * 7

                # Record mass error
                this_feature_info[4] = ppm_error(mz_in_list, mz)
                
                # Record intensity
                this_feature_info[5] = avg_int
                
                # Check the CCS value
                if ccs is not None:
                    ccs_error = percent_error(ccs, sample_info[ii, 2])
                    this_feature_info[6] = ccs_error
                else:
                    this_feature_info[6] = np.nan
                
                # Add to all info
                if (PARAMS[14] == 0) or (ccs is None) or (ccs_error < PARAMS[5]):
                    feature_info.append(this_feature_info)

        # Check for other close masses
        mass_list.remove(mz_in_list)  # Don't want to reconsider same mass
        mz_in_list = in_list(mass_list, mz)
    
    return feature_info


#%% FTICR-MS

# Remove noise/contamination from MS data
def downselect_fticr(library_masses, adduct_masses, fticr_features, e=MAX_FTICR_ERROR):
    library_masses = np.sort(library_masses)  # For faster searching
    delete_ind = range(fticr_features.shape[0])
    for z in adduct_masses:
        library_masses += z  # Adduct masses for library
        
        # Find features that do NOT match any of these adduct masses
        temp_delete_ind = []
        for i in range(fticr_features.shape[0]):
            mass_in_lib = find_closest_sorted(library_masses, fticr_features[i, 1])
            if not np.isclose(mass_in_lib, fticr_features[i, 1], atol=e):
                temp_delete_ind.append(i)
        
        # Only keep what is in both lists
        delete_ind = list(set(delete_ind).intersection(temp_delete_ind))
        
        library_masses -= z  # Reset library masses
        
    return np.delete(fticr_features, delete_ind, axis=0)


# Read in all FTICR-MS data
def read_fticr_data(fname = '../data/FTICR_%s.csv', stop=13):
    fticr_pos = np.loadtxt(fname % 'Pos_Features', delimiter=',', skiprows=1)
    fticr_neg = np.loadtxt(fname % 'Neg_Features', delimiter=',', skiprows=1)
    all_formulas = list(np.loadtxt(fname % 'Formulas', dtype=str))
    neg_formulas = all_formulas[1: all_formulas.index('Positive')]
    pos_formulas = all_formulas[all_formulas.index('Positive') + 1:]
    return fticr_pos[:, :stop], fticr_neg[:, :stop], pos_formulas, neg_formulas


# Calculate and apply a shift to all mass values in MS data
def shift_fticr_mass(fticr_pos, fticr_neg, library_masses):
    adducts = [1.0078246, 22.9898, -1.0078246, 34.968853]  # ['H', 'Na', 'De', 'Cl']
    pos_shift, neg_shift = np.nan, np.nan
    if PARAMS[13] == 1:
        fticr_pos, fticr_neg = np.copy(fticr_pos), np.copy(fticr_neg)
        
        pos_shift = calc_mass_shift(adducts[:2], PARAMS[8], library_masses, fticr_pos[:, 1])
        fticr_pos[:, 1] += pos_shift / 1000000.
        
        neg_shift = calc_mass_shift(adducts[2:], PARAMS[8], library_masses, fticr_neg[:, 1])
        fticr_neg[:, 1] += neg_shift / 1000000.

    return fticr_pos, fticr_neg, [pos_shift, neg_shift]


# Remove features seen in the blank or below the minimum intensity
def _process_fticr_data(features, sample, blank):
    
    # Only keep features seen in the sample and not in the blank
    ind = np.where(np.logical_and(features[:, sample] >= PARAMS[9],
                                  features[:, blank] < PARAMS[10]))[0]
    features = features[ind, :]
    
    return features


# Remove possible noise/contamination from MS data
# Note: assumes only one replicate
def process_fticr_data(fticr_pos, fticr_neg, sample, blank):
    
    fticr_pos = _process_fticr_data(np.copy(fticr_pos), sample, blank)
    fticr_neg = _process_fticr_data(np.copy(fticr_neg), sample, blank)
    
    return fticr_pos, fticr_neg


# Get all FTICR-MS evidence for a given library entry's mass
def get_fticr_evidence(mz, full_list):
    mass_list = np.copy(full_list[:, 1])
    feature_info = []
    mz_in_list = in_list(mass_list, mz, e=PARAMS[8], sorted_list=True)
    while mz_in_list is not None:
        
        # Record match. Note - assumes only one possible match (unlike IMS-MS)
        ind = list(full_list[:, 1]).index(mz_in_list)
        
        # Start scoring
        this_feature_info = [0] * 7

        # Record mass error
        this_feature_info[4] = ppm_error(mz_in_list, mz)
        
        # Record intensity
        this_feature_info[5] = full_list[ind, 1]
        
        # Add to all info
        feature_info.append(this_feature_info)
        
        # Check for any other close matches
        mass_list = np.delete(mass_list, np.where(mass_list == mz_in_list))
        mz_in_list = in_list(mass_list, mz, e=PARAMS[8], sorted_list=True)
    
    return feature_info


#%% Feature matching!

# For all compounds in a mix, find all features from the experimental
# data that matches something in the suspect library
def feature_matching(mix_num, imsms_pos, imsms_neg, fticr_pos, fticr_neg,
                     pos_formulas, neg_formulas, adducts, cpd_masses,
                     cpd_formulas, cpd_ccs_vals):
    all_mixes = np.array(['Mix', 'Cpd No', 'Instr', 'Adduct', 'MassE', 'Int',
                          'Other'])
    
    fticr_pos, fticr_neg, fticr_shifts = shift_fticr_mass(fticr_pos, fticr_neg, cpd_masses)
    imsms_pos, imsms_neg, shifts = shift_imsms_mass(imsms_pos, imsms_neg, cpd_masses)
    shifts.extend(fticr_shifts)  # Want to return for tracking reasons
    
    for mix in range(mix_num):
        
        # Get & process IMS-MS data for this mix
        usecols = get_usecols(mix)  # Columns associated with this mix
        pos_feat, neg_feat = imsms_pos[:, usecols], imsms_neg[:, usecols]
        pos_inf, pos_int, neg_inf, neg_int, pos_masses, neg_masses = \
        downselect_imsms_features(pos_feat, neg_feat)
    
        # Get & process FT-ICR data for this mix
        s = mix + 2  # Column specific to this mix
        blank = mix_num + 2
        fticr_pos_ref, fticr_neg_ref = process_fticr_data(fticr_pos, fticr_neg,
                                                          s, blank)

        # Begin cycling through each library entry and its possible matches
        for cpd_no in range(cpd_masses.shape[0]):
    
            # Initalize compound information          
            ccs_vals = cpd_ccs_vals[cpd_no, :]
            formula = cpd_formulas[cpd_no]
            mass = cpd_masses[cpd_no]
    
            # Initialize list to collect all evidence
            feature_info = []
            
            # Record FT-ICR mass matches
            full_list = fticr_pos_ref
            formulas = pos_formulas
            for j in range(len(adducts)):
    
                # Switch to negative mode variables
                if j == len(adducts) / 2:
                    full_list = fticr_neg_ref
                    formulas = neg_formulas
                
                mz = mass + adducts[j]
                formula = fmod.formula_to_string(fmod.formula_split(formula))
                this_adduct_info = get_fticr_evidence(mz, full_list)
                
                # Label this evidenxe as being from this compound, FT-ICR,
                # and adduct
                if len(this_adduct_info) > 0:
                    this_adduct_info = np.array(this_adduct_info)
                    this_adduct_info[:, 0] = mix
                    this_adduct_info[:, 1] = cpd_no + 2
                    this_adduct_info[:, 2] = 0 # This is FT-ICR
                    this_adduct_info[:, 3] = j # Record which adduct
                    
                    # Record if isotopic sig seen
                    if formula in formulas:
                        this_adduct_info[:, 6] = 1
                    
                    this_adduct_info = [list(x) for x in this_adduct_info]
                    
                    # Append to all feature evidence
                    feature_info.extend(this_adduct_info)
    
            # Record IMS-MS mass/CCS matches
            sample_info = pos_inf
            sample_masses = pos_masses
            sample_int = pos_int
            for j in range(len(adducts) - 1):
                
                # Adducts[2+] are negative mode
                if j == len(adducts) / 2:
                    sample_info = neg_inf
                    sample_masses = neg_masses
                    sample_int = neg_int
    
                # Set mz, ccs, and init mass list for this compound + adduct
                mz = mass + adducts[j]
                ccs = ccs_vals[j]
                this_adduct_info = get_imsms_evidence(mz, ccs, sample_masses[:],
                                                      sample_info, sample_int)
                
                # Label this evidenxe as being from this compound, FT-ICR,
                # and adduct
                if len(this_adduct_info) > 0:
                    this_adduct_info = np.array(this_adduct_info)
                    this_adduct_info[:, 0] = mix
                    this_adduct_info[:, 1] = cpd_no + 2
                    this_adduct_info[:, 2] = 1 # This is IMS-MS
                    this_adduct_info[:, 3] = j # Record which adduct
                    this_adduct_info = [list(x) for x in this_adduct_info]
                    
                    # Append to all feature evidence
                    feature_info.extend(this_adduct_info)
            
            # Done. Save all info for this compound/mix combo in overall results
            if len(feature_info) > 0:
                all_mixes = np.vstack((all_mixes, np.array(feature_info)))
    # Cut out labels in first row and convert to float64 array
    all_mixes = np.array(all_mixes[1:, :], dtype=np.float64)

    return all_mixes, shifts